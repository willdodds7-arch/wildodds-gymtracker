"""Regenerates catalogue_index65.xlsx — fixture for DefaultProgramXlsxParserTest.

A miniature default_programs.xlsx with a 65-row Program Index plus 65 one-sheet-per-program
sheets ("N-ProgN"). Each program sheet deliberately has NO RPE/%1RM data cells, so the parser
can only derive category + RPE/%1RM tracking from the INDEX — which is exactly what the
no-hard-cap fix must guarantee for every one of the 65 entries.

Run:  python app/src/test/resources/fixtures/make_catalogue_fixture.py
Requires: openpyxl. (DefaultProgramXlsxParser handles openpyxl's absolute rels targets, so no
post-processing is needed — unlike the SmartXlsxParser fixture.)
"""
import os
import openpyxl

N = 65

# Goal -> category is resolved by the parser; we cycle real goals and assert the mapping.
GOALS = ["Hypertrophy", "Powerlifting", "Powerbuilding", "Strength",
         "Athletic", "GPP", "Conditioning", "Bodyweight"]
# Intensity strings the parser reads for RPE / %1RM tracking.
INTENSITIES = ["RPE", "%1RM", "RPE/%1RM", "Linear", ""]

wb = openpyxl.Workbook()

# ── Program Index sheet ──────────────────────────────────────────────────────
idx = wb.active
idx.title = "Program Index"
idx["A1"] = "Mini Catalogue — Index"
headers = ["#", "Program Name", "Creator", "Level", "Goal", "Weeks", "Days/Wk",
           "Sheet Tab", "Intensity System"]
for c, h in enumerate(headers, start=1):
    idx.cell(row=2, column=c, value=h)

for i in range(1, N + 1):
    goal = GOALS[(i - 1) % len(GOALS)]
    intensity = INTENSITIES[(i - 1) % len(INTENSITIES)]
    tab = f"{i}-Prog{i}"
    r = i + 2  # data starts on row 3
    idx.cell(row=r, column=1, value=i)
    idx.cell(row=r, column=2, value=f"Prog{i}")
    idx.cell(row=r, column=3, value="Example Author")
    idx.cell(row=r, column=4, value="Intermediate")
    idx.cell(row=r, column=5, value=goal)
    idx.cell(row=r, column=6, value="4")
    idx.cell(row=r, column=7, value=3)
    idx.cell(row=r, column=8, value=tab)
    idx.cell(row=r, column=9, value=intensity)

# ── One sheet per program (N-ProgN) ──────────────────────────────────────────
for i in range(1, N + 1):
    ws = wb.create_sheet(f"{i}-Prog{i}")
    ws["A1"] = f"Prog{i}"
    # Row 2 metadata (label:/value pairs, mirroring the real catalogue).
    ws["A2"] = "Creator:"; ws["B2"] = "Example Author"
    ws["C2"] = "Level:";   ws["D2"] = "Intermediate"
    ws["E2"] = "Goal:";    ws["F2"] = GOALS[(i - 1) % len(GOALS)]
    ws["G2"] = "Total Weeks:"; ws["H2"] = "4"
    ws["A3"] = "Example program — fixture only."
    # Row 4 headers: Week | Day/Focus | Exercise | Sets | Reps | RPE | %1RM | Notes
    for c, h in enumerate(["Week", "Day / Focus", "Exercise", "Sets", "Reps",
                           "RPE", "%1RM / TM%", "Notes"], start=1):
        ws.cell(row=4, column=c, value=h)
    # Row 5: one exercise, NO RPE / %1RM cells -> tracking can only come from the index.
    ws["A5"] = "Week 1"; ws["B5"] = "Day 1 - Full Body"
    ws["C5"] = "Squat";  ws["D5"] = "3"; ws["E5"] = "5"
    ws["A6"] = None; ws["B6"] = None
    ws["C6"] = "Bench Press"; ws["D6"] = "3"; ws["E6"] = "5"

out = os.path.join(os.path.dirname(__file__), "catalogue_index65.xlsx")
wb.save(out)
print("wrote", out, "with", N, "programs")
