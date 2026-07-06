"""Generates app/src/main/assets/bulk_template.xlsx — the phased bulk-upload template
for Wild Odds Gym Tracker. Run with: python build_bulk_template.py"""
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

# Brand colours
DARK_BLUE = "1E3A5F"
ACCENT    = "60B4E8"
LIGHT_BLUE= "B8DDF5"
BG        = "F0F8FF"

HEADERS = ["Program", "Phase", "Phase Name", "Phase Weeks", "Day",
           "Exercise", "Sets", "Reps", "RPE", "%1RM", "Rest", "Superset", "Notes"]

header_fill = PatternFill("solid", fgColor=DARK_BLUE)
header_font = Font(bold=True, color="FFFFFF", size=11)
phase_fill  = PatternFill("solid", fgColor=LIGHT_BLUE)
title_font  = Font(bold=True, color=DARK_BLUE, size=16)
sub_font    = Font(italic=True, color="40566F", size=10)
thin        = Side(style="thin", color="C9DDEC")
border      = Border(left=thin, right=thin, top=thin, bottom=thin)
center      = Alignment(horizontal="center", vertical="center")
left        = Alignment(horizontal="left", vertical="center", wrap_text=True)

def style_header(ws, row=1):
    for c, h in enumerate(HEADERS, start=1):
        cell = ws.cell(row=row, column=c, value=h)
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = center
        cell.border = border
    widths = [22, 7, 18, 12, 16, 26, 6, 9, 7, 8, 8, 9, 30]
    for i, w in enumerate(widths, start=1):
        ws.column_dimensions[get_column_letter(i)].width = w
    ws.freeze_panes = "A2"

def write_row(ws, row, values):
    for c, v in enumerate(values, start=1):
        cell = ws.cell(row=row, column=c, value=v)
        cell.border = border
        cell.alignment = left if c in (1, 3, 5, 6, 13) else center

wb = Workbook()

# ───────────────────────── Instructions sheet ─────────────────────────
ins = wb.active
ins.title = "Instructions"
ins.sheet_view.showGridLines = False
ins.column_dimensions["A"].width = 4
ins.column_dimensions["B"].width = 100

lines = [
    ("Wild Odds Gym Tracker — Bulk Upload Template", title_font),
    ("Turn program screenshots into a ready-to-import spreadsheet.", sub_font),
    ("", None),
    ("HOW TO USE", Font(bold=True, color=DARK_BLUE, size=12)),
    ("1. Give your program screenshots to Claude along with this file.", None),
    ("2. Ask Claude to fill in the 'Programs' sheet, one row per exercise.", None),
    ("3. Save the file, then in the app: Library → Import from Excel → Upload, or", None),
    ("   the Bulk Import tile. Every program in the sheet is added to your Library.", None),
    ("", None),
    ("COLUMNS", Font(bold=True, color=DARK_BLUE, size=12)),
    ("Program      The program name. Multiple programs can share one sheet —", None),
    ("             just change this value when a new program starts.", None),
    ("Phase        Phase number (1, 2, 3 …). Leave at 1 for single-block programs.", None),
    ("Phase Name   e.g. Accumulation, Intensification, Peak, Deload. Optional.", None),
    ("Phase Weeks  How many weeks this phase runs. The week you fill in is", None),
    ("             repeated for that many weeks (log progressive weights live).", None),
    ("Day          Day/session name, e.g. 'Push A', 'Lower', 'Arms'. Each distinct", None),
    ("             Day under a Phase becomes one session.", None),
    ("Exercise     Movement name. One row per exercise.", None),
    ("Sets         Number of working sets (a number).", None),
    ("Reps         Rep target, e.g. 8, 8-12, AMRAP.", None),
    ("RPE          Target RPE, e.g. 8 or 7-8. Optional — enables RPE tracking.", None),
    ("%1RM         Target % of 1-rep-max or TM%, e.g. 75%. Optional — enables %1RM.", None),
    ("Rest         Rest between sets, e.g. 90s, 2min. Optional.", None),
    ("Superset     A letter (A, B …) to link exercises into a superset. Optional.", None),
    ("Notes        Any cue or instruction. Optional.", None),
    ("", None),
    ("TIPS", Font(bold=True, color=DARK_BLUE, size=12)),
    ("• You only need to fill ONE week per phase — the app repeats it.", None),
    ("• Program / Phase / Phase Name / Phase Weeks / Day can be left blank on", None),
    ("  continuation rows; the value above carries down (see the Example sheet).", None),
    ("• Phases are optional. A program with only Phase 1 looks exactly like a", None),
    ("  normal program in the app.", None),
    ("• See the 'Example' sheet for a complete 2-phase program.", None),
]
r = 1
for text, font in lines:
    cell = ins.cell(row=r, column=2, value=text)
    if font:
        cell.font = font
    r += 1

# ───────────────────────── Programs (blank template) ─────────────────────────
prog = wb.create_sheet("Programs")
prog.sheet_view.showGridLines = False
style_header(prog)
# a couple of starter rows so the structure is obvious
starter = [
    ["My Program", 1, "Accumulation", 4, "Push", "Barbell Bench Press", 4, "6-8", "8", "", "2min", "", "Top set then back-offs"],
    ["",           "", "",            "",  "",    "Overhead Press",     3, "8-10","8", "", "90s", "", ""],
    ["",           "", "",            "",  "Pull","Deadlift",           3, "5",   "8", "", "3min","", ""],
]
for i, row in enumerate(starter, start=2):
    write_row(prog, i, row)

# ───────────────────────── Example sheet ─────────────────────────
ex = wb.create_sheet("Example")
ex.sheet_view.showGridLines = False
style_header(ex)
example = [
    # Program, Phase, Phase Name, Phase Weeks, Day, Exercise, Sets, Reps, RPE, %1RM, Rest, Superset, Notes
    ["PHUL Power", 1, "Power", 4, "Upper Power", "Barbell Bench Press", 4, "3-5", "8", "", "3min", "", "Heavy"],
    ["", "", "", "", "", "Bent-Over Row", 4, "3-5", "8", "", "3min", "", ""],
    ["", "", "", "", "", "Overhead Press", 3, "5-8", "7", "", "2min", "", ""],
    ["", "", "", "", "Lower Power", "Squat", 4, "3-5", "8", "", "3min", "", ""],
    ["", "", "", "", "", "Deadlift", 3, "3-5", "8", "", "3min", "", ""],
    ["", "", "", "", "", "Leg Press", 3, "10-15", "", "", "2min", "", ""],
    ["", 2, "Hypertrophy", 4, "Upper Hypertrophy", "Incline Dumbbell Press", 4, "8-12", "9", "", "90s", "A", ""],
    ["", "", "", "", "", "Cable Fly", 3, "12-15", "9", "", "60s", "A", "Superset with press"],
    ["", "", "", "", "", "Lat Pulldown", 4, "10-12", "9", "", "90s", "", ""],
    ["", "", "", "", "Lower Hypertrophy", "Hack Squat", 4, "10-12", "9", "", "90s", "", ""],
    ["", "", "", "", "", "Romanian Deadlift", 3, "10-12", "9", "", "90s", "", ""],
    ["", "", "", "", "", "Standing Calf Raise", 4, "12-15", "", "", "60s", "", ""],
    # second program in the same sheet
    ["Arm Blast", 1, "", 6, "Arms", "Barbell Curl", 4, "8-12", "", "", "75s", "A", ""],
    ["", "", "", "", "", "Skull Crusher", 4, "8-12", "", "", "75s", "A", ""],
    ["", "", "", "", "", "Hammer Curl", 3, "12-15", "", "", "60s", "", ""],
]
prev_phase = None
for i, row in enumerate(example, start=2):
    write_row(ex, i, row)
    # light shade rows where a new phase starts
    if row[1] not in ("", None) and row[1] != prev_phase:
        for c in range(1, len(HEADERS) + 1):
            ex.cell(row=i, column=c).fill = phase_fill
        prev_phase = row[1]
    elif row[0] not in ("", None):
        prev_phase = None

wb.save("app/src/main/assets/bulk_template.xlsx")
print("wrote app/src/main/assets/bulk_template.xlsx")
