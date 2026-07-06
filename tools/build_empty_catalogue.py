#!/usr/bin/env python3
"""
Generate an EMPTY default_programs.xlsx — a catalogue with no programs, ready for the next update.

It contains only a "Program Index" sheet documenting the per-program layout the parser reads,
including the v18 browse metadata (Coach / Coach Bio / Days per Week / Split / Style). Because there
are no per-program sheets, DefaultProgramXlsxParser.parse() returns an empty list, so the in-app
"Browse Programs" screen shows its empty state.

Per-program sheet layout (for when programs are added back):
  Row 1: Program title (col A)
  Row 2: label/value metadata pairs, e.g.
         Coach: | Jane Doe | Days per Week: | 4 | Split: | Upper/Lower | Style: | Hypertrophy |
                | Total Weeks: | 8 | Goal: | Hypertrophy | Coach Bio: | <short bio>
  Row 3: Description (col A, free text)
  Row 4: Headers  Week | Day/Focus | Exercise | Sets | Reps | RPE | %1RM/TM% | Notes
  Row 5+: Data rows (Week/Day carry forward when blank)

Index sheet columns (row 2 headers): # | Name | Coach | Level | Goal | Weeks | Days | Split | Style | SheetTab | Intensity System
"""
import openpyxl
from openpyxl.styles import Font, PatternFill

OUT = "app/src/main/assets/default_programs.xlsx"

DARK = "1E3A5F"

wb = openpyxl.Workbook()
ws = wb.active
ws.title = "📋 Program Index"

ws["A1"] = "Wild Odds — Program Index"
ws["A1"].font = Font(bold=True, size=14, color="FFFFFF")
ws["A1"].fill = PatternFill("solid", fgColor=DARK)

# Column order matches tools/add_program.py (Sheet Tab = col 8, Intensity = col 9) so the authoring
# tool keeps appending correctly; Coach/Split/Style are appended after for the v18 browse metadata.
headers = ["#", "Name", "Creator", "Level", "Goal", "Weeks", "Days",
           "Sheet Tab", "Intensity System", "Coach", "Split", "Style"]
for i, h in enumerate(headers, start=1):
    c = ws.cell(row=2, column=i, value=h)
    c.font = Font(bold=True, color="FFFFFF")
    c.fill = PatternFill("solid", fgColor=DARK)

# No data rows — the catalogue is intentionally empty for this release.
ws.cell(row=3, column=1, value="(no programs yet — added in a future update)")

wb.save(OUT)
print("wrote empty catalogue ->", OUT)
