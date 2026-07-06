#!/usr/bin/env python3
"""Append a coach program to default_programs.xlsx from a structured JSON spec.

Mirrors the catalogue's per-program sheet layout EXACTLY (the layout
DefaultProgramXlsxParser reads):

    Row 1            A: program title
    Row 2            A:"Creator:" B:<creator> C:"Level:" D:<level>
                     E:"Goal:" F:<goal> G:"Total Weeks:" H:<weeks>
    Row 3            A: description
    Row 4 (headers)  Week | Day / Focus | Exercise | Sets | Reps | RPE | %1RM / TM% | Notes
    Row 5+           per-day exercise rows; Week (A) + Day (B) carry forward (blank on
                     continuation rows), matching the merged-cell style the parser expects.

…and appends a matching Program Index row (#, Name, Creator, Level, Goal, Weeks, Days/Wk,
Sheet Tab, Intensity System).

COLUMN MAPPING — the spec accepts the official single-program template vocabulary
(Exercise/Sets/Reps/Weight/%1RM/RPE/Rest/Superset/Notes), but the catalogue sheet has only
8 columns. Exercise/Sets/Reps/RPE/%1RM/Notes map 1:1; Weight, Rest and Superset have no column
in the catalogue and are folded into Notes (e.g. "… | Rest 180s | Superset A | @100kg") so no
information is lost. Catalogue programs are a single representative week (the parser does not
auto-repeat); `weeks` is informational and lands in the index + row-2 metadata.

USAGE
    python tools/add_program.py spec.json                 # append to the bundled asset (backs it up)
    python tools/add_program.py spec.json --output out.xlsx   # write a copy, leave the asset alone
    python tools/add_program.py spec.json --dry-run       # validate only, write nothing

See tools/example_program_spec.json for the spec format.
"""
import argparse
import json
import os
import re
import shutil
import sys

import openpyxl

DEFAULT_WORKBOOK = os.path.join("app", "src", "main", "assets", "default_programs.xlsx")
HEADERS = ["Week", "Day / Focus", "Exercise", "Sets", "Reps", "RPE", "%1RM / TM%", "Notes"]


def slugify_title(name: str) -> str:
    """Sheet names can't contain []:*?/\\ and are capped at 31 chars by Excel."""
    cleaned = re.sub(r'[\[\]:\*\?/\\]', "", name).strip()
    return cleaned[:24]  # leave room for the "N-" prefix within Excel's 31-char limit


def next_program_number(wb) -> int:
    nums = []
    for name in wb.sheetnames:
        m = re.match(r'^(\d+)-', name)
        if m:
            nums.append(int(m.group(1)))
    return (max(nums) + 1) if nums else 1


def find_index_sheet(wb):
    for name in wb.sheetnames:
        if "index" in name.lower() or name.startswith("\U0001f4cb"):
            return wb[name]
    raise SystemExit("ERROR: no Program Index sheet found in the workbook.")


def folded_notes(ex: dict) -> str:
    parts = []
    if ex.get("notes"):
        parts.append(str(ex["notes"]).strip())
    if ex.get("weight"):
        parts.append(f"@{str(ex['weight']).strip()}")
    if ex.get("rest"):
        parts.append(f"Rest {str(ex['rest']).strip()}")
    if ex.get("superset"):
        parts.append(f"Superset {str(ex['superset']).strip()}")
    return " | ".join(parts)


def append_program(wb, spec: dict) -> str:
    number = next_program_number(wb)
    sheet_tab = f"{number}-{slugify_title(spec['name'])}"

    ws = wb.create_sheet(sheet_tab)
    ws["A1"] = spec["name"]
    days_per_week = spec.get("days_per_week", len(spec["days"]))
    ws["A2"] = "Creator:";       ws["B2"] = spec.get("creator", "")
    ws["C2"] = "Level:";         ws["D2"] = spec.get("level", "")
    ws["E2"] = "Goal:";          ws["F2"] = spec.get("goal", "")
    ws["G2"] = "Total Weeks:";   ws["H2"] = str(spec.get("weeks", ""))
    # v18 browse metadata — read by DefaultProgramXlsxParser (Coach maps from Creator; these add tags).
    ws["I2"] = "Days per Week:"; ws["J2"] = str(days_per_week)
    ws["K2"] = "Split:";         ws["L2"] = spec.get("split", "")
    ws["M2"] = "Style:";         ws["N2"] = spec.get("style", "")
    ws["O2"] = "Coach Bio:";     ws["P2"] = spec.get("coach_bio", "")
    ws["A3"] = spec.get("description", "")
    for c, h in enumerate(HEADERS, start=1):
        ws.cell(row=4, column=c, value=h)

    row = 5
    for day in spec["days"]:
        first = True
        for ex in day["exercises"]:
            if first:
                ws.cell(row=row, column=1, value="Week 1")
                ws.cell(row=row, column=2, value=day["name"])
                first = False
            ws.cell(row=row, column=3, value=str(ex["exercise"]))
            ws.cell(row=row, column=4, value=str(ex.get("sets", "")))
            ws.cell(row=row, column=5, value=str(ex.get("reps", "")))
            ws.cell(row=row, column=6, value=str(ex.get("rpe", "")))
            ws.cell(row=row, column=7, value=str(ex.get("pct1rm", "")))
            ws.cell(row=row, column=8, value=folded_notes(ex))
            row += 1

    # Program Index row.
    idx = find_index_sheet(wb)
    r = idx.max_row + 1
    idx.cell(row=r, column=1, value=number)
    idx.cell(row=r, column=2, value=spec["name"])
    idx.cell(row=r, column=3, value=spec.get("creator", ""))
    idx.cell(row=r, column=4, value=spec.get("level", ""))
    idx.cell(row=r, column=5, value=spec.get("goal", ""))
    idx.cell(row=r, column=6, value=str(spec.get("weeks", "")))
    idx.cell(row=r, column=7, value=str(days_per_week))
    idx.cell(row=r, column=8, value=sheet_tab)
    idx.cell(row=r, column=9, value=spec.get("intensity", ""))
    idx.cell(row=r, column=10, value=spec.get("creator", ""))   # Coach
    idx.cell(row=r, column=11, value=spec.get("split", ""))
    idx.cell(row=r, column=12, value=spec.get("style", ""))
    return sheet_tab


def validate(path: str, sheet_tab: str) -> None:
    """Re-open the saved workbook and confirm the program + index row are present & well-formed."""
    wb = openpyxl.load_workbook(path, data_only=True)
    assert sheet_tab in wb.sheetnames, f"sheet {sheet_tab!r} missing after save"
    ws = wb[sheet_tab]
    assert [ws.cell(row=4, column=c).value for c in range(1, 9)] == HEADERS, "header row 4 malformed"
    assert ws.cell(row=5, column=3).value, "no exercise data on row 5"
    idx = find_index_sheet(wb)
    tabs = [idx.cell(row=r, column=8).value for r in range(3, idx.max_row + 1)]
    assert sheet_tab in tabs, "index row missing"
    print(f"  validation OK — {sheet_tab} present, headers intact, index row added "
          f"({len(wb.sheetnames)} sheets total)")


def main():
    ap = argparse.ArgumentParser(description="Append a coach program to default_programs.xlsx")
    ap.add_argument("spec", help="path to the program spec JSON")
    ap.add_argument("--workbook", default=DEFAULT_WORKBOOK, help="source workbook (default: bundled asset)")
    ap.add_argument("--output", help="write to this file instead of overwriting the workbook")
    ap.add_argument("--dry-run", action="store_true", help="validate against a temp file; write nothing permanent")
    args = ap.parse_args()

    with open(args.spec, encoding="utf-8") as f:
        spec = json.load(f)
    for key in ("name", "goal", "days"):
        if key not in spec:
            sys.exit(f"ERROR: spec is missing required field '{key}'")

    wb = openpyxl.load_workbook(args.workbook)
    sheet_tab = append_program(wb, spec)

    if args.dry_run:
        tmp = args.workbook + ".dryrun.tmp.xlsx"
        wb.save(tmp)
        validate(tmp, sheet_tab)
        os.remove(tmp)
        print(f"DRY RUN — would add {sheet_tab!r}; nothing written.")
        return

    out = args.output or args.workbook
    if not args.output:
        backup = args.workbook + ".bak"
        shutil.copyfile(args.workbook, backup)
        print(f"  backed up original -> {backup}")
    wb.save(out)
    validate(out, sheet_tab)
    print(f"DONE — added {sheet_tab!r} to {out}")


if __name__ == "__main__":
    main()
